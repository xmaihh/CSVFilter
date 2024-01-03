import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re


class BeautifyExcel:
    def __init__(self, excel_filepath):
        self.excel_filepath = excel_filepath

    def _beautify_excel(self, worksheet):
        # 对 Excel 文件进行美化操作
        # 可以设置样式、添加图表、调整列宽等

        # 设置标题行的字体和对齐方式
        title_font = Font(bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.font = title_font
            cell.alignment = title_alignment

        # 设置数据行的对齐方式
        data_alignment = Alignment(horizontal="left", vertical="center")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment

        # 自动调整列宽
        dims = {}  # 设置一个字典用于保存列宽数据
        # 遍历表格数据，获取自适应列宽数据
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.7 * len(
                        re.findall("([\u4e00-\u9fa5])", str(cell.value))
                    ) + len(str(cell.value))
                    dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
        for col, value in dims.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2

    def read_all_worksheets(self):
        # 打开Excel文件
        workbook = openpyxl.load_workbook(self.excel_filepath)
        worksheet_names = workbook.sheetnames
        for worksheet_name in worksheet_names:
            worksheet = workbook[worksheet_name]
            # 美化Excel文件
            self._beautify_excel(worksheet)
        try:
            workbook.save(self.excel_filepath)
            print(f"Beautify Excel. Saved to {self.excel_filepath}")
            return self.excel_filepath
        except PermissionError:
            raise Exception(
                "Permission denied. Please close the Excel file and try again."
            )
