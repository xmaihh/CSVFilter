import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re
from csv_toolbox.lib_base.base_filter import BaseFilter
from csv_toolbox.lib_base.constants import EXCEL_BEAUTIFIER_PREFIX, XLSX_EXTENSION


class ExcelBeautifier(BaseFilter):
    def __init__(self, input_path):
        super().__init__(input_path, EXCEL_BEAUTIFIER_PREFIX, XLSX_EXTENSION)

    def _beautify_excel(self):
        # Load the Excel file
        workbook = openpyxl.load_workbook(self.input_path)
        # Get all worksheet names
        worksheet_names = workbook.sheetnames
        # Loop through all worksheets
        for worksheet_name in worksheet_names:
            # Get the current worksheet
            worksheet = workbook[worksheet_name]
            # Set the font and alignment of title row
            title_font = Font(bold=True)
            title_alignment = Alignment(horizontal="center", vertical="center")
            for cell in worksheet[1]:# The first row is the title row
                cell.font = title_font
                cell.alignment = title_alignment

            # Set the alignment of data rows
            data_alignment = Alignment(horizontal="left", vertical="center")
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment

            # Automatically adjust column width
            dims = {}  # A dictionary to store column width data
            # Traverse the table data to obtain adaptive column width data
            for row in worksheet.rows:
                for cell in row:
                    if cell.value:
                        # Traverse the entire table and compare the length of all cells in the column to find the longest cell
                        # When comparing cell text, Chinese characters should be recognized as a length of 1.7, and English characters should be recognized as a length of 1.
                        # Here, the text length can be directly added to the number of Chinese characters recognized by the regular expression re.findall('([\u4e00-\u9fa5])', cell.value) can recognize most Chinese characters
                        cell_len = 0.7 * len(
                            re.findall("([\u4e00-\u9fa5])", str(cell.value))
                        ) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
            for col, value in dims.items():
                # Set the column width. get_column_letter is used to get the letter column corresponding to the digital column number. The final value of +2 is used to adjust the final effect.
                worksheet.column_dimensions[get_column_letter(col)].width = value + 2
        try:
            # Save to Excel file
            workbook.save(self.output_path)
        except PermissionError:
            raise Exception(
                "Permission denied. Please close the Excel file and try again."
            )

    def beautify(self):
        self._beautify_excel()
        self.save_to()
