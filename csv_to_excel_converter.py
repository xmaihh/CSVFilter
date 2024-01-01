import os
import chardet
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import openpyxl
from ..helpers import const
import re

"""
__init__(self, csv_filepath):初始化方法,接收CSV文件的路径作为参数,并初始化其他实例变量。
_get_excel_filepath(self)：内部方法，用于构建输出的 Excel 文件路径。您可以在此方法中根据需求自定义 Excel 文件的名称和保存位置，并返回 Excel 文件路径。
_convert_csv_to_dataframe(self)：内部方法，用于读取 CSV 文件并将数据转换为 DataFrame。返回 DataFrame。
_beautify_excel(self, worksheet)：内部方法，用于对 Excel 文件进行美化操作。您可以在此方法中设置样式、添加图表、调整列宽等。
convert_csv_to_excel(self)：主要方法，用于将 CSV 文件转换为 Excel 文件。在该方法中，创建 Excel 文件和工作表，将 CSV 数据写入 Excel 文件，并调用 _beautify_excel 方法对 Excel 文件进行美化。最后，保存 Excel 文件
"""


class CsvToExcelConverter:
    def __init__(self, csv_filepath):
        self.csv_filepath = csv_filepath
        self.excel_filepath = self._get_excel_filepath()
        self.file_encoding = self._get_file_encoding()

    def _get_excel_filepath(self):
        # 获取 CSV 文件所在的目录路径
        csv_dir = os.path.dirname(self.csv_filepath)
        # 获取 CSV 文件的文件名（不包含扩展名）
        csv_filename = os.path.splitext(os.path.basename(self.csv_filepath))[0]
        # 构建 Excel 文件的文件名
        excel_filename = (
            f"{const.EXCEL_FILE_PREFIX}{csv_filename}{const.EXCEL_FILE_SUFFIX}"
        )
        # 拼接 Excel 文件的完整路径
        excel_filepath = os.path.join(csv_dir, excel_filename)
        print(excel_filepath)
        return excel_filepath

    def _get_file_encoding(self):
        # 检测文件编码
        # 获取检测到的编码格式
        with open(self.csv_filepath, "rb") as f:
            result = chardet.detect(f.read())
            return result["encoding"]

    def _convert_csv_to_dataframe(self):
        # 读取 CSV 文件并将数据转换为 DataFrame
        # 返回 DataFrame
        try:
            df = pd.read_csv(self.csv_filepath, encoding=self.file_encoding)
            return df
        except FileNotFoundError:
            raise Exception("CSV file not found.")
        except pd.errors.EmptyDataError:
            raise Exception("CSV file is empty.")
        except pd.errors.ParserError:
            raise Exception("Error parsing CSV file.")

    def _beautify_excel(self, worksheet):
        # 对 Excel 文件进行美化操作
        # 可以设置样式、添加图表、调整列宽等

        # 设置标题行的字体和对齐方式
        title_font = Font(bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.font = title_font
            cell.alignment = title_alignment

        # 设置数据行的对齐方式
        data_alignment = Alignment(horizontal="left", vertical="center")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment

        # 自动调整列宽
        dims = {}  # 设置一个字典用于保存列宽数据
        # 遍历表格数据，获取自适应列宽数据
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.7 * len(
                        re.findall("([\u4e00-\u9fa5])", str(cell.value))
                    ) + len(str(cell.value))
                    dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
        for col, value in dims.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2

    def convert_csv_to_excel(self):
        # 将 CSV 文件转换为 Excel 文件
        # 创建 Excel 文件和工作表
        # 将 CSV 数据写入 Excel 文件
        # 调用 _beautify_excel 方法对 Excel 文件进行美化
        # 保存 Excel 文件

        # 转换 CSV 文件为 Pandas DataFrame 对象
        df = self._convert_csv_to_dataframe()
        # 转换成Excel文件
        df.to_excel(self.excel_filepath, sheet_name="row_data")
        # 打开Excel文件
        workbook = openpyxl.load_workbook(self.excel_filepath)
        worksheet = workbook.active

        # 美化 Excel 文件
        self._beautify_excel(worksheet)
        # 保存 Excel 文件
        try:
            workbook.save(self.excel_filepath)
            print(f"Converted CSV to Excel. Saved to {self.excel_filepath}")
            return self.excel_filepath
        except PermissionError:
            raise Exception(
                "Permission denied. Please close the Excel file and try again."
            )